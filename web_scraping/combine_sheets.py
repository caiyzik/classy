import openpyxl as xl

def search(match, data):
    match = str(match).lower()
    for i in range(len(data)):
        r = str(data[i][0].value).lower()
        print(r)
        if(r.find(str(match).lower()) != -1):
            return True
        else:
            return False

wb = xl.load_workbook('combined.xlsx')
oscar = wb.get_sheet_by_name('Oscar')
uc3m = wb.get_sheet_by_name('UC3M')
trans_num = oscar['A1':'A{0}'.format(oscar.max_row)]
trans_name = oscar['B1':'B{0}'.format(oscar.max_row)]
color1 = xl.styles.colors.Color(rgb='00FF0000')
color2 = xl.styles.colors.Color(rgb='0023FF00')
fill1= xl.styles.fills.PatternFill(patternType='solid', fgColor=color1)
fill2 = xl.styles.fills.PatternFill(patternType='solid', fgColor=color2)

count1 = 0
count2 = 0
for i in range(uc3m.max_row):
    course_num = uc3m.cell(row=i+1, column=1)
    course_name = uc3m.cell(row=i+1, column=2)
    print(course_name)
    if(search(course_num.value, trans_num)):
        course_num.fill = fill1
        count1+=1
    if(search(course_name.value, trans_name)):
        course_name.fill = fill2
        count2+=1
wb.save('combined.xlsx')
print(count1)
print(count2)
