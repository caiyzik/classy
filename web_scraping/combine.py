import openpyxl as xl

#Takes: start cell, end cell, and sheet you want to copy from.
def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)
 
    return rangeSelected

#Paste range
#Paste data from copyRange into template sheet
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow+1,1):
        countCol = 0
        for j in range(startCol,endCol+1,1):
            
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1

wb = xl.load_workbook('UC3M_transfer.xlsx')
oscar = wb.get_sheet_by_name('Oscar')

for i in range(oscar.max_row):
    course = oscar.cell(row=i+1, column=4)
    if(course.value == 'And'):
	#move this cell + 4 to previous row column 8
        ran = copyRange(4,i+1,7,i+1, oscar)
        pasteRange(8,i,11,i,oscar, ran)
    	#next_row = oscar.cell(row=i+2, column=4)
        #if(next_row.value == 'Or'):
            

wb.save('NewOscar.xlsx')
