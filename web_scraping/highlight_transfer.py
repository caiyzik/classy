import openpyxl as xl

wb = xl.load_workbook('transferable.xlsx')
oscar = wb.get_sheet_by_name('Oscar')
uc3m = wb.get_sheet_by_name('UC3M')

oscar_nums = oscar['A1':'A{0}'.format(oscar.max_row)]
oscar_names = oscar['B1':'B{0}'.format(oscar.max_row)]

red = xl.styles.colors.Color(rgb='00FF0000')
green = xl.styles.colors.Color(rgb='0000FF00')

fill1= xl.styles.fills.PatternFill(patternType='solid', fgColor=red)
fill2 = xl.styles.fills.PatternFill(patternType='solid', fgColor=green)

match = "14360"

def transfer(match, data):
    for cell in oscar_nums:
        course_num = str(cell[0].value)
        print("This is the OSCAR course num: {0}".format(course_num))
        if (course_num.find(match) != -1):
            print("Found a match!")
            return True

for i in range(uc3m.max_row):
    course_num = uc3m.cell(row=i+1, column=1)
    print(course_num.value)
    if transfer(str(course_num.value), oscar_nums):
        course_num.fill = fill2
wb.save('transferable.xlsx')

